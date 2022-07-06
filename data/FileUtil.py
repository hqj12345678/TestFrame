#index,title,playsource,category,result,TestData,tester
import openpyxl
from config import setting
import configparser
class FileUtil():
    def __init__(self):
        self.path=setting.TestCase_Path
        self.book=openpyxl.open(self.path)
    def ReadData(self,sheetname):
        datas=[]
        #sheet=self.book.get_sheet_by_name(sheetname)
        sheet=self.book[sheetname]
        row_num=sheet.max_row
        for row in range(2,row_num):
            data={}
            data["index"]=sheet.cell(row=row,column=1).value
            data["title"]=sheet.cell(row=row,column=2).value
            data["playsource"]=sheet.cell(row=row,column=3).value
            data["category"]=sheet.cell(row=row,column=4).value
            datas.append(data)
        return  datas
    def WriteData(self,row,writes):
        sheet=self.book.active
        sheet.cell(row=row,column=5,value=writes["result"])
        sheet.cell(row=row,column=6,value=writes["testdata"])
        sheet.cell(row=row,column=7,value=writes["tester"])
    def SaveBook(self):
        self.book.save(self.path)





